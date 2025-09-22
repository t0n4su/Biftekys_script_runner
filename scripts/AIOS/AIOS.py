import pandas as pd
import numpy as np
import re
import os
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def clean_application_number(num):
    """Standardize application number format (1-* or VOD*)"""
    if pd.isna(num) or num is None:
        return ""
    
    num = str(num).strip()
    
    if 'LB_VODAFONE' in num:
        num_parts = re.findall(r'\d+', num)
        if num_parts:
            num = max(num_parts, key=len)
    
    if re.match(r'^1-', num):
        return num
    if re.match(r'^VOD', num, re.IGNORECASE):
        return 'VOD' + re.sub(r'^VOD', '', num, flags=re.IGNORECASE)
    if re.match(r'^\d{10,}', num):
        return '1-' + num
    if re.match(r'^\d+$', num):
        return '1-' + num
    
    if not num.startswith('VOD'):
        return 'VOD' + num
    return num

def split_name(full_name):
    """Split full name into first and last name"""
    if pd.isna(full_name) or full_name is None:
        return ("", "")
    
    full_name = str(full_name).strip()
    parts = full_name.split()
    
    if len(parts) == 0:
        return ("", "")
    if len(parts) == 1:
        return (parts[0], "")
    
    return (' '.join(parts[:-1]), parts[-1])

def extract_mobile_number(mobile_data):
    """Extract mobile number from various formats"""
    if pd.isna(mobile_data) or mobile_data is None:
        return ""
    
    mobile_str = str(mobile_data)
    numbers = re.findall(r'\d+', mobile_str)
    if numbers:
        longest_num = max(numbers, key=len)
        if len(longest_num) == 10 and longest_num.startswith('69'):
            return longest_num
        return longest_num
    
    return ""

def clean_duplicate_values(value):
    """Clean duplicate values in a cell"""
    if pd.isna(value) or value is None:
        return ""
    
    value_str = str(value)
    if ',' in value_str:
        parts = [part.strip() for part in value_str.split(',')]
        unique_parts = []
        for part in parts:
            if part and part not in unique_parts and part != 'nan':
                unique_parts.append(part)
        return ', '.join(unique_parts) if unique_parts else ""
    
    return value_str

def process_files(input_files, output_dir):
    """
    Main function to process Excel files and generate result.xlsx
    """
    all_df = None
    aios_dp_df = None
    aios_mob_df = None
    aios_one_net_df = None
    
    # Αποθήκευση όλων των αριθμών αιτήσεων από AIOS αρχεία
    aios_app_numbers = set()
    
    for file_path in input_files:
        file_name = os.path.basename(file_path).lower()
        
        try:
            if 'all' in file_name:
                df = pd.read_excel(file_path)
                all_df = df.iloc[:, [11, 12, 14, 17]].copy()
                all_df.columns = ['Επώνυμο', 'Όνομα', 'Αριθμός Αίτησης', 'Κινητό']
                print(f"Processed ALL file: {file_name}")
                
            elif 'aios_dp' in file_name or 'dp_aios' in file_name:
                df = pd.read_excel(file_path)
                aios_dp_df = df.iloc[:, [0, 10, 11]].copy()
                aios_dp_df.columns = ['Αριθμός Αίτησης', 'Όνομα', 'Επώνυμο']
                # Προσθήκη αριθμών αιτήσεων από AIOS αρχεία
                aios_app_numbers.update(aios_dp_df['Αριθμός Αίτησης'].apply(clean_application_number))
                print(f"Processed AIOS_DP file: {file_name}")
                
            elif 'aios_mob' in file_name or 'mob_aios' in file_name:
                df = pd.read_excel(file_path)
                aios_mob_df = df.iloc[:, [1, 5]].copy()
                aios_mob_df.columns = ['Αριθμός Αίτησης', 'Ονοματεπώνυμο']
                aios_app_numbers.update(aios_mob_df['Αριθμός Αίτησης'].apply(clean_application_number))
                print(f"Processed AIOS_MOB file: {file_name}")
                
            elif 'aios_one' in file_name or 'one_aios' in file_name or 'one_net' in file_name:
                df = pd.read_excel(file_path)
                aios_one_net_df = df.iloc[:, [1, 6]].copy()
                aios_one_net_df.columns = ['Αριθμός Αίτησης', 'Ονοματεπώνυμο']
                aios_app_numbers.update(aios_one_net_df['Αριθμός Αίτησης'].apply(clean_application_number))
                print(f"Processed AIOS_ONE_NET file: {file_name}")
                
        except Exception as e:
            print(f"Error processing {file_name}: {str(e)}")
            continue
    
    # Δημιουργία τελικού DataFrame με όλες τις αιτήσεις από ALL
    if all_df is None:
        print("No ALL file found!")
        return False
        
    # Καθαρισμός αριθμών αιτήσεων και κινητών
    all_df['Αριθμός Αίτησης'] = all_df['Αριθμός Αίτησης'].apply(clean_application_number)
    all_df['Κινητό'] = all_df['Κινητό'].apply(extract_mobile_number)
    
    # Προσθήκη στήλης κατάστασης
    all_df['Κατάσταση'] = all_df['Αριθμός Αίτησης'].apply(
        lambda x: 'ΕΝΕΡΓΟΠΟΙΗΜΕΝΗ' if x in aios_app_numbers else 'ΜΗ ΕΝΕΡΓΟΠΟΙΗΜΕΝΗ'
    )
    
    # Αφαίρεση διπλοεγγραφών
    result_df = all_df.drop_duplicates(subset=['Αριθμός Αίτησης'])
    
    # Ταξινόμηση: πρώτα οι ΜΗ ΕΝΕΡΓΟΠΟΙΗΜΕΝΕΣ, μετά οι ΕΝΕΡΓΟΠΟΙΗΜΕΝΕΣ
    result_df = result_df.sort_values(by='Κατάσταση', ascending=False)
    
    # Αποθήκευση αποτελεσμάτων χωρίς μορφοποίηση πρώτα
    output_path = os.path.join(output_dir, 'result.xlsx')
    result_df.to_excel(output_path, index=False)
    
    # Τώρα ανοίγουμε το αρχείο για να προσθέσουμε μορφοποίηση
    wb = Workbook()
    ws = wb.active
    
    # Ορισμός χρωμάτων
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    
    # Ανάγνωση του αρχείου που δημιουργήσαμε
    formatted_df = pd.read_excel(output_path)
    
    # Προσθήκη επικεφαλίδων
    headers = list(formatted_df.columns)
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Προσθήκη δεδομένων και εφαρμογή μορφοποίησης
    for row_idx, row_data in formatted_df.iterrows():
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx+2, column=col_idx, value=value)
        
        # Εφαρμογή χρωματικής μορφοποίησης στη στήλη "Κατάσταση"
        status_cell = ws.cell(row=row_idx+2, column=len(headers))
        if row_data['Κατάσταση'] == 'ΜΗ ΕΝΕΡΓΟΠΟΙΗΜΕΝΗ':
            status_cell.fill = red_fill
        else:
            status_cell.fill = green_fill
    
    # Αποθήκευση του αρχείου με μορφοποίηση
    wb.save(output_path)
    print(f"Final results saved to: {output_path}")
    
    # Εμφάνιση στατιστικών
    total_applications = len(result_df)
    energized = len(result_df[result_df['Κατάσταση'] == 'ΕΝΕΡΓΟΠΟΙΗΜΕΝΗ'])
    not_energized = len(result_df[result_df['Κατάσταση'] == 'ΜΗ ΕΝΕΡΓΟΠΟΙΗΜΕΝΗ'])
    
    print(f"\n📋Συνολικές αιτήσεις: {total_applications}")
    print(f"✅Ενεργοποιημένες: {energized}")
    print(f"❌Μη ενεργοποιημένες: {not_energized}")
    
    return True

def main():
    input_files = sys.argv[1:]
    output_dir = os.environ.get('OUTPUT_DIR', os.getcwd())
    
    if not input_files:
        print("No input files provided. Please drag and drop files onto the GUI.")
        return
    
    print(f"Processing {len(input_files)} files...")
    print(f"Output directory: {output_dir}")
    
    success = process_files(input_files, output_dir)
    
    if success:
        print("Processing completed successfully!")
    else:
        print("Processing failed!")
        sys.exit(1)

if __name__ == "__main__":
    main()